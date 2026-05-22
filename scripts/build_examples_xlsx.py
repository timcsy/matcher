"""產生 examples/{teacher-class,study-group}/roster*.xlsx。"""
from pathlib import Path
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent


def build_teacher_class():
    wb = Workbook()
    ws = wb.active
    ws.title = "報名表"
    ws.append(["id", "姓名", "專業科目", "年資"])
    rows = [
        ("T01", "王老師", "國文", 8),
        ("T02", "林老師", "數學", 6),
        ("T03", "陳老師", "英文", 4),
        ("T04", "李老師", "國文", 3),
        ("T05", "張老師", "數學", 10),
        ("T06", "黃老師", "自然", 5),
        ("T07", "周老師", "英文", 7),
        ("T08", "吳老師", "社會", 4),
        ("T09", "鄭老師", "國文", 12),
        ("T10", "蔡老師", "自然", 3),
    ]
    for r in rows:
        ws.append(r)
    out = ROOT / "examples" / "teacher-class" / "roster.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


def build_study_group_single():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "姓名", "年級"])
    students = [
        ("S01", "小明", 5), ("S02", "小華", 4), ("S03", "小美", 6),
        ("S04", "小強", 5), ("S05", "小芳", 4), ("S06", "小傑", 6),
        ("S07", "小婷", 5), ("S08", "小宇", 4), ("S09", "小薇", 6),
    ]
    for s in students:
        ws.append(s)
    out = ROOT / "examples" / "study-group" / "roster.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


def build_study_group_multi():
    wb = Workbook()
    # 移除預設工作表
    wb.remove(wb.active)
    # 報名表
    ws = wb.create_sheet("報名表")
    ws.append(["id", "姓名", "年級"])
    students = [
        ("S01", "小明", 5), ("S02", "小華", 4), ("S03", "小美", 6),
        ("S04", "小強", 5), ("S05", "小芳", 4), ("S06", "小傑", 6),
        ("S07", "小婷", 5), ("S08", "小宇", 4), ("S09", "小薇", 6),
    ]
    for s in students:
        ws.append(s)
    # 說明
    ws2 = wb.create_sheet("說明")
    ws2.append(["這是說明工作表"])
    # 範例
    ws3 = wb.create_sheet("範例")
    ws3.append(["範例工作表"])

    out = ROOT / "examples" / "study-group" / "roster-multi.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    build_teacher_class()
    build_study_group_single()
    build_study_group_multi()

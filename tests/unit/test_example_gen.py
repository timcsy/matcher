"""Feature 016 US2：動態範例試算表產生。"""

from __future__ import annotations

import csv
import io

from matcher.template_loader import TemplateRegistry
from matcher.web.example_gen import participant_example_bytes, target_example_bytes

REG = TemplateRegistry()


def _csv_rows(b: bytes):
    return list(csv.reader(io.StringIO(b.decode("utf-8-sig"))))


def test_target_example_headers_match_template():
    rows = _csv_rows(target_example_bytes(REG.get("teacher-class"), "csv"))
    headers = rows[0]
    assert headers == ["編號", "容量", "班級名稱", "班級需要的科目清單", "班級特色"]


def test_participant_example_headers_match_template():
    rows = _csv_rows(participant_example_bytes(REG.get("teacher-class"), "csv"))
    headers = rows[0]
    assert headers[0] == "編號"
    assert "老師姓名" in headers
    assert "老師專業科目" in headers


def test_example_has_format_hint_row():
    rows = _csv_rows(target_example_bytes(REG.get("teacher-class"), "csv"))
    hint = "".join(rows[1])
    assert "數字" in hint  # 容量
    assert "分號" in hint  # list_str 欄位提示


def test_xlsx_example_generates_and_reads_back():
    from openpyxl import load_workbook
    data = target_example_bytes(REG.get("teacher-class"), "xlsx")
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[:2] == ["編號", "容量"]


def test_example_syncs_with_template_attrs():
    """study-group 對象屬性（分組名稱/研習主題/最低年級）→ 範例表頭含之。"""
    rows = _csv_rows(target_example_bytes(REG.get("study-group"), "csv"))
    headers = rows[0]
    assert "編號" in headers and "容量" in headers
    # study-group 對象屬性的中文顯示名稱應出現
    assert any("組" in h or "主題" in h or "年級" in h for h in headers)

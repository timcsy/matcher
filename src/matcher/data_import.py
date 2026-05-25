"""CSV / Excel 名單匯入。

依模板宣告的 attributes schema（含 aliases）對齊欄位，產出與 YAML 路徑等價的 Roster。
targets 由旁檔 `<stem>.targets.yaml` 提供（CSV/Excel 為平表格式，不適合表達 roles+targets 兩段）。

研究決策見 specs/003-data-import/research.md。
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from matcher.errors import (
    EmptyRoster,
    RosterColumnMismatch,
    RosterDecodeError,
    RosterSheetAmbiguous,
    RosterTypeError,
)
from matcher.roster import Role, Roster
from matcher.template import AttributeDecl, Template

ENCODINGS_TO_TRY = ("utf-8", "utf-8-sig", "cp950")


# ── 編碼偵測 ─────────────────────────────────────────────────────────


UTF8_BOM = b"\xef\xbb\xbf"


def detect_csv_encoding(raw_bytes: bytes) -> tuple[str, str]:
    """依序嘗試 UTF-8 → UTF-8-SIG → CP950。

    BOM 檢測優先：若 bytes 開頭為 UTF-8 BOM，直接視為 utf-8-sig。
    回傳 (encoding, decoded_text)。三輪皆失敗 → RosterDecodeError。
    """
    if raw_bytes.startswith(UTF8_BOM):
        try:
            return "utf-8-sig", raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            pass

    for enc in ENCODINGS_TO_TRY:
        try:
            return enc, raw_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    raise RosterDecodeError(
        f"無法解碼 CSV 檔案。\n"
        f"細節：已嘗試編碼 {', '.join(ENCODINGS_TO_TRY)} 皆失敗。\n"
        f"建議：請以 Excel 或文字編輯器另存為 UTF-8 編碼後再試。"
    )


# ── 表頭對齊 ─────────────────────────────────────────────────────────


def _is_ascii(s: str) -> bool:
    return all(ord(c) < 128 for c in s)


def resolve_header(name: str, decls: tuple) -> Optional[AttributeDecl]:
    """將 CSV 表頭對齊到模板宣告的 AttributeDecl。

    對齊邏輯：strip → 精確 key → ASCII 不分大小寫 key → 精確 alias → ASCII 不分大小寫 alias。
    """
    normalized = name.strip()

    for decl in decls:
        if normalized == decl.key:
            return decl
        if _is_ascii(normalized) and _is_ascii(decl.key) and normalized.lower() == decl.key.lower():
            return decl

    for decl in decls:
        # 別稱清單 + 顯示名稱本身（自動視為 alias，讓使用者不必重複填寫）
        candidate_aliases = list(decl.aliases) + ([decl.description] if decl.description else [])
        for alias in candidate_aliases:
            if normalized == alias:
                return decl
            if _is_ascii(normalized) and _is_ascii(alias) and normalized.lower() == alias.lower():
                return decl

    return None


# ── 型別轉換 ─────────────────────────────────────────────────────────


def coerce_value(raw, decl: AttributeDecl, *, row_num: int) -> object:
    """依模板宣告的型別轉換 CSV/Excel 儲存格值。失敗 → RosterTypeError。"""
    if decl.type == "str":
        if raw is None:
            return ""
        return str(raw).strip()

    if decl.type == "int":
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            raise RosterTypeError(
                f"第 {row_num} 列、欄位 `{decl.key}` 缺值。\n"
                f"細節：模板宣告為 int 型別，但值為空。\n"
                f"建議：填入整數或調整模板宣告。"
            )
        if isinstance(raw, int) and not isinstance(raw, bool):
            return int(raw)
        try:
            return int(str(raw).strip())
        except (ValueError, TypeError):
            raise RosterTypeError(
                f"第 {row_num} 列、欄位 `{decl.key}` 型別轉換失敗。\n"
                f"細節：值「{raw}」無法解析為整數（int）。\n"
                f"建議：請改為阿拉伯數字（如 8），或調整模板宣告為 str。"
            )

    if decl.type == "list_str":
        if raw is None:
            return []
        if not isinstance(raw, str):
            raise RosterTypeError(
                f"第 {row_num} 列、欄位 `{decl.key}` 型別轉換失敗。\n"
                f"細節：模板宣告為 list_str，但值非字串：{type(raw).__name__}。\n"
                f"建議：以分號分隔的字串填入（例：G1; G2; G3）。"
            )
        s = raw.strip()
        if not s:
            return []
        return [part.strip() for part in s.split(";") if part.strip()]

    raise RosterTypeError(f"欄位 `{decl.key}` 的型別宣告 `{decl.type}` 不支援")


# ── preferences 別名（由模板 ui_fields 推導） ────────────────────────


def _find_id_header(headers: list) -> Optional[str]:
    """找出可選的 `id` 欄位（不分大小寫，亦接受別名「編號」「id」）。"""
    for h in headers:
        if h is None:
            continue
        normalized = str(h).strip()
        if normalized.lower() == "id" or normalized == "編號":
            return h
    return None


def _preferences_aliases(template: Template) -> set[str]:
    out = {"preferences"}
    for u in template.ui_fields:
        if u.key == "preferences":
            out.add(u.label)
    return out


# ── 共用 helpers ─────────────────────────────────────────────────────


def _resolve_headers(
    headers: list[str],
    template: Template,
) -> tuple[dict, set]:
    """回傳 (header_text → AttributeDecl, 偵測到的 preferences 表頭集合)。"""
    role_decls = template.attributes.roles
    pref_aliases = _preferences_aliases(template)

    header_resolved: dict = {}
    pref_headers: set = set()
    seen_keys: dict = {}
    for h in headers:
        if h is None or not str(h).strip():
            continue
        normalized = str(h).strip()
        if normalized in pref_aliases:
            pref_headers.add(h)
            continue
        decl = resolve_header(normalized, role_decls)
        if decl is None:
            continue
        if decl.key in seen_keys:
            raise RosterColumnMismatch(
                f"表頭重複：欄位 `{decl.key}` 由兩個表頭對齊。\n"
                f"細節：第一個是「{seen_keys[decl.key]}」，第二個是「{h}」。\n"
                f"建議：移除重複欄位。"
            )
        seen_keys[decl.key] = h
        header_resolved[h] = decl

    required_role_keys = [d.key for d in role_decls if d.required]
    resolved_keys = {d.key for d in header_resolved.values()}
    missing = [k for k in required_role_keys if k not in resolved_keys]
    if missing:
        miss_info = []
        for k in missing:
            decl = next((d for d in role_decls if d.key == k), None)
            aliases = list(decl.aliases) if decl else []
            alias_text = "、".join(aliases) if aliases else "（無）"
            miss_info.append(f"`{k}`（可用別名：{alias_text}）")
        raise RosterColumnMismatch(
            f"表頭缺少模板必填欄位。\n"
            f"細節：缺漏 {'; '.join(miss_info)}。\n"
            f"建議：在表頭新增上述欄位（key 或別名皆可），或調整模板宣告。"
        )

    return header_resolved, pref_headers


def _build_roles(
    rows: list,
    header_resolved: dict,
    pref_headers: set,
    id_header: Optional[str] = None,
) -> tuple:
    if not rows:
        raise EmptyRoster("名單為空：只有表頭、無資料列")

    seen_ids: set = set()
    roles_list = []
    for idx, row in enumerate(rows, start=1):
        if id_header is not None and row.get(id_header):
            role_id = str(row[id_header]).strip()
        else:
            role_id = f"R{idx:03d}"
        from matcher.errors import DuplicateIdentity
        if role_id in seen_ids:
            raise DuplicateIdentity(f"名單有重複身分：角色 id `{role_id}` 在第 {idx + 1} 列重複出現")
        seen_ids.add(role_id)
        attrs: dict = {}
        prefs_raw = None

        for header_text, value in row.items():
            if header_text in pref_headers:
                prefs_raw = value
                continue
            decl = header_resolved.get(header_text)
            if decl is None:
                continue
            attrs[decl.key] = coerce_value(value, decl, row_num=idx + 1)

        if isinstance(prefs_raw, str):
            s = prefs_raw.strip()
            prefs = tuple(p.strip() for p in s.split(";") if p.strip()) if s else ()
        elif isinstance(prefs_raw, list):
            prefs = tuple(str(p).strip() for p in prefs_raw if str(p).strip())
        else:
            prefs = ()

        roles_list.append(Role(id=role_id, attributes=attrs, preferences=prefs))

    return tuple(roles_list)


def _load_targets(path: Path, template: Template) -> tuple:
    """取得 targets：優先讀旁檔 `<stem>.targets.yaml`；若不存在則使用 template.default_targets。"""
    sidecar = path.parent / f"{path.stem}.targets.yaml"
    if sidecar.exists():
        import yaml
        from matcher.roster import Target
        with sidecar.open("r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        raw_targets = data.get("targets") or []
        if not raw_targets:
            raise RosterColumnMismatch(f"旁檔 {sidecar.name} 缺少 `targets` 段或為空。")
        target_ids: list[str] = []
        targets = []
        for t in raw_targets:
            tid = t["id"]
            if tid in target_ids:
                from matcher.errors import DuplicateIdentity
                raise DuplicateIdentity(f"旁檔中對象 id `{tid}` 出現多次")
            target_ids.append(tid)
            cap = int(t.get("capacity", 1))
            if cap < 1:
                raise ValueError(f"對象 `{tid}` 的容量 {cap} 無效；容量必須 ≥ 1")
            targets.append(Target(id=tid, capacity=cap, attributes=dict(t.get("attributes", {}))))
        return tuple(targets)

    # Fallback：使用模板的 default_targets
    if template.default_targets:
        return template.default_targets

    raise RosterColumnMismatch(
        f"找不到對象（targets）來源：旁檔 {sidecar.name} 不存在、模板 `{template.id}` 也未宣告 default_targets。\n"
        f"細節：CSV/Excel 路徑下，targets 須由旁檔提供，或由模板 default_targets 自含。\n"
        f"建議：建立 {sidecar.name}、或讓模板宣告 default_targets、或改用 --roster <yaml>。"
    )


# ── CSV 匯入 ─────────────────────────────────────────────────────────


def load_roster_csv(path: str | Path, template: Template) -> tuple[Roster, dict]:
    """從 CSV 載入 Roster；targets 由旁檔提供。"""
    import csv
    import io

    p = Path(path)
    raw_bytes = p.read_bytes()
    encoding, text = detect_csv_encoding(raw_bytes)

    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    header_resolved, pref_headers = _resolve_headers(headers, template)
    id_header = _find_id_header(headers)

    rows = list(reader)
    roles = _build_roles(rows, header_resolved, pref_headers, id_header=id_header)
    targets = _load_targets(p, template)

    metadata = {
        "source_type": "csv",
        "encoding": encoding,
        "sheet_name": None,
        "row_count": len(rows),
        "file_basename": p.name,
    }
    return Roster(roles=roles, targets=targets), metadata


# ── Excel 匯入 ─────────────────────────────────────────────────────


def load_roster_xlsx(
    path: str | Path,
    template: Template,
    sheet: Optional[str] = None,
) -> tuple[Roster, dict]:
    """從 .xlsx 載入 Roster；targets 由旁檔提供。"""
    from openpyxl import load_workbook

    p = Path(path)
    wb = load_workbook(p, read_only=True, data_only=True)

    sheet_names = wb.sheetnames
    if sheet is None:
        if len(sheet_names) > 1:
            wb.close()
            raise RosterSheetAmbiguous(
                f"Excel 檔含多張工作表，未指定 --sheet。\n"
                f"細節：可用工作表：{'、'.join(repr(s) for s in sheet_names)}。\n"
                f"建議：以 --sheet <name> 指定要匯入的工作表。"
            )
        chosen = sheet_names[0]
    else:
        if sheet not in sheet_names:
            wb.close()
            raise RosterSheetAmbiguous(
                f"指定的工作表 `{sheet}` 不存在。\n"
                f"細節：可用工作表：{'、'.join(repr(s) for s in sheet_names)}。\n"
                f"建議：以 --sheet <實際工作表名> 指定。"
            )
        chosen = sheet

    ws = wb[chosen]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise EmptyRoster("名單為空：Excel 工作表沒有任何列")

    headers = [str(h) if h is not None else "" for h in header_row]
    header_resolved, pref_headers = _resolve_headers(headers, template)
    id_header = _find_id_header(headers)

    # 建立 header → 欄位 index 對應（包含 preferences 與可選的 id）
    header_index: dict = {}
    for i, h in enumerate(headers):
        if h and (h in header_resolved or h in pref_headers or h == id_header):
            header_index[h] = i

    row_dicts = []
    for row in rows_iter:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        d = {h: (row[i] if i < len(row) else None) for h, i in header_index.items()}
        row_dicts.append(d)

    wb.close()

    roles = _build_roles(row_dicts, header_resolved, pref_headers, id_header=id_header)
    targets = _load_targets(p, template)

    metadata = {
        "source_type": "xlsx",
        "encoding": None,
        "sheet_name": chosen,
        "row_count": len(row_dicts),
        "file_basename": p.name,
    }
    return Roster(roles=roles, targets=targets), metadata
